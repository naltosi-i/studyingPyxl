import openpyxl
import pathlib
from datetime import datetime

file_path = 'studyingPyxlBook1.xlsx'

'''
wb = openpyxl.load_workbook('studyingPyxlBook1.xlsx') # Excelファイルを読み込むメソッド
# pathlib.Path('test').mkdir(exist_ok=True)
# wb.save('test/studyingPyxlBook3.xlsx') # 変数に格納されたExcelファイルを保存するメソッド
'''

wb = openpyxl.load_workbook(file_path)
ws = wb.worksheets[0]

''' selecting cells and substitutting value for cells
c1 = ws['A1']
c2 = ws['A2']
c3 = ws['A3']
c4 = ws['A4']

c1.value = 'test'
c2.value = 'Hello world.'
c3.value = 101
c4.value = '101'
'''

c5 = ws['A5']
c6 = ws['A6']

now = datetime.now()
c5.value = now
c6.value = '=CONCAT(A1:A2)'

wb.save(file_path)